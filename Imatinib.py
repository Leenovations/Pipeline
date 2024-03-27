import os
import time
import argparse
import pandas as pd
import numpy as np
from collections import defaultdict
from datetime import datetime
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl import load_workbook
#import sys
#import smtplib
#import math
#import glob
#from fpdf import FPDF
#----------------------------------------------------------------------------------------#
Start_time = time.time()
dt = datetime.now()
#----------------------------------------------------------------------------------------#
parser = argparse.ArgumentParser(description='Pipeline Usage')
args = parser.parse_args()
#----------------------------------------------------------------------------------------#
Sample = pd.read_csv('SampleSheet.txt', sep='\t', header=None)
Name = Sample.iloc[0,0]
R1 = Sample.iloc[0,1]
R2 = Sample.iloc[0,2]
#----------------------------------------------------------------------------------------#
BATCH = {}
with open(f"{Name}.batch.config", "r") as batch:
    for line in batch:
        line = line.strip()
        splitted = line.split("=")
        Key = splitted[0]
        Value = splitted[1]
        BATCH[Key] = Value
#----------------------------------------------------------------------------------------#        
def PreQC(name, r1, r2):
    if os.path.isdir('00.PreQC'):
        pass
    else:
        command = 'mkdir 00.PreQC'
    os.system(command)

    command =f'fastqc -o 00.PreQC\
            -t {BATCH["CPU"]}\
            {r1}\
            {r2}'
    os.system(command)
#----------------------------------------------------------------------------------------#
def Trimming(name, r1, r2):
    if os.path.isdir('02.Trimmed'):
        pass
    else:
        command = 'mkdir 02.Trimmed'
    os.system(command)

    command = f'/media/src/Tools/TrimGalore-0.6.10/trim_galore \
                --paired --gzip \
                -j {BATCH["CPU"]} \
                -o 02.Trimmed --basename {name} \
                {r1} {r2}'
    os.system(command)
#----------------------------------------------------------------------------------------#
def PostQC(name, r1, r2):
    if os.path.isdir('01.PostQC'):
        pass
    else:
        command = 'mkdir 01.PostQC'
    os.system(command)

    command = f'fastqc -o 01.PostQC \
                -t {BATCH["CPU"]} \
                02.Trimmed/{name}_val_1.fq.gz \
                02.Trimmed/{name}_val_2.fq.gz'
    os.system(command)
#----------------------------------------------------------------------------------------#
def Refindex():
    if os.path.isdir(f'/Bioinformatics/01.Reference/hg{BATCH["Ref.ver"].split("g")[1]}/hg{BATCH["Ref.ver"].split("g")[1]}_gencode/'):
        pass
    else:
        command = f'/media/src/Tools/STAR-2.7.11b/source/STAR \
                    --runThreadN {BATCH["CPU"]} \
                    --runMode genomeGenerate \
                    --genomeDir /Bioinformatics/01.Reference/hg{BATCH["Ref.ver"].split("g")[1]}/hg{BATCH["Ref.ver"].split("g")[1]}_gencode/ \
                    --sjdbOverhang 100 \
                    --sjdbGTFfile /Bioinformatics/01.Reference/hg{BATCH["Ref.ver"].split("g")[1]}/gencode.v{BATCH["Ref.ver"].split("g")[1]}.annotation.gtf \
                    --genomeFastaFiles /Bioinformatics/01.Reference/b37/b37.fa'
    os.system(command)
#----------------------------------------------------------------------------------------#
def STAR(name):
    command = f'/media/src/Tools/STAR-2.7.11b/source/STAR \
                --runThreadN {BATCH["CPU"]} \
                --genomeDir /Bioinformatics/01.Reference/hg{BATCH["Ref.ver"].split("g")[1]}/hg{BATCH["Ref.ver"].split("g")[1]}_gencode/ \
                --genomeLoad NoSharedMemory \
                --readFilesIn 02.Trimmed/{name}_val_1.fq.gz 02.Trimmed/{name}_val_2.fq.gz --readFilesCommand zcat \
                --outStd Log --outSAMtype BAM Unsorted --outSAMunmapped Within --outBAMcompression 0 \
                --outFilterMultimapNmax 50 --peOverlapNbasesMin 10 --alignSplicedMateMapLminOverLmate 0.5 \
                --alignSJstitchMismatchNmax 5 -1 5 5 --chimSegmentMin 10 --chimOutType WithinBAM HardClip \
                --chimJunctionOverhangMin 10 --chimScoreDropMax 30 --chimScoreJunctionNonGTAG 0 --chimScoreSeparation 1 \
                --chimSegmentReadGapMax 3 --chimMultimapNmax 50 \
                --outFileNamePrefix 03.Output/{name}_'
    os.system(command)
#----------------------------------------------------------------------------------------#
def QC(name):
    if os.path.isdir('04.QC/'):
        pass
    else:
        command = 'mkdir 04.QC'
    os.system(command)

    command = f'samtools sort 03.Output/{name}_Aligned.out.bam -o 03.Output/{name}_Sorted.out.bam'
    os.system(command)

    command = f'samtools index 03.Output/{name}_Sorted.out.bam'
    os.system(command)

    command = f'mv 03.Output/{name}_Log.final.out 04.QC/'
    os.system(command)

    command = f'samtools stats 03.Output/{name}_Aligned.out.bam > 04.QC/{name}.stats'
    os.system(command)

    command = f'samtools stats -t /Bioinformatics/01.Reference/Panel/rna.gene.bed 03.Output/{name}_Sorted.out.bam > 04.QC/{name}.Ontarget.stats'
    os.system(command)

    command = f'samtools flagstat 03.Output/{name}_Sorted.out.bam > 04.QC/{name}.Depth.txt'
    os.system(command)

    command = f'samtools depth 03.Output/{name}_Sorted.out.bam > 04.QC/{name}.Total.Depth.txt'
    os.system(command)

    command = f'samtools depth -b /Bioinformatics/01.Reference/Panel/BCR-ABL1.bed \
                03.Output/{name}_Sorted.out.bam > 04.QC/{name}.Target.Depth.txt'
    os.system(command)
#----------------------------------------------------------------------------------------#
def Somatic(name):
    # command = f'java -jar /Bioinformatics/00.Tools/picard/build/libs/picard.jar \
    #             AddOrReplaceReadGroups \
    #             TMP_DIR=03.Output/TEMP_PICARD \
    #             VALIDATION_STRINGENCY=SILENT \
    #             SO=coordinate \
    #             I=03.Output/{name}_Sorted.out.bam \
    #             O=03.Output/{name}.read.bam \
    #             RGID={name} \
    #             RGLB={name} \
    #             RGPL=Illumina \
    #             RGPU={name} \
    #             RGSM={name} CREATE_INDEX=true'
    # os.system(command)

    # command = f'java -jar /Bioinformatics/00.Tools/picard/build/libs/picard.jar \
    #             MarkDuplicates \
    #             I=03.Output/{name}.read.bam \
    #             O=03.Output/{name}.dedup.bam \
    #             M=03.Output/{name}.dedup.metrics \
    #             VALIDATION_STRINGENCY=SILENT \
    #             REMOVE_DUPLICATES=true'
    # os.system(command)

    # command = f'samtools index 03.Output/{name}.dedup.bam'
    # os.system(command)

    command = f'java -jar /media/src/Tools/GATK-3.7/GenomeAnalysisTK.jar \
                -T SplitNCigarReads \
                -R /Bioinformatics/01.Reference/b37/b37.fa \
                -L /Bioinformatics/01.Reference/Panel/BCR-ABL1.bed \
                -I 03.Output/{name}.dedup.bam \
                -rf ReassignOneMappingQuality -RMQF 255 -RMQT 60 -U ALLOW_N_CIGAR_READS \
                -o 03.Output/{name}.SplitNCigar.bam'
    os.system(command)

    command = f'java -jar /media/src/Tools/GATK-3.7/GenomeAnalysisTK.jar \
                -T RealignerTargetCreator \
                -nt {BATCH["CPU"]} \
                -R /Bioinformatics/01.Reference/b37/b37.fa \
                -L /Bioinformatics/01.Reference/Panel/BCR-ABL1.bed \
                -known /Bioinformatics/01.Reference/b37/mills.b37.vcf \
                -I 03.Output/{name}.SplitNCigar.bam \
                -o 03.Output/{name}.intervals'
    os.system(command)

    command = f'java -jar /media/src/Tools/GATK-3.7/GenomeAnalysisTK.jar \
                -T IndelRealigner \
                -R /Bioinformatics/01.Reference/b37/b37.fa \
                -L /Bioinformatics/01.Reference/Panel/BCR-ABL1.bed \
                -known /Bioinformatics/01.Reference/b37/mills.b37.vcf \
                --maxReadsForRealignment 10000000 --maxReadsInMemory 10000000 \
                -targetIntervals 03.Output/{name}.intervals \
                -I 03.Output/{name}.SplitNCigar.bam \
                -o 03.Output/{name}.indel.bam'
    os.system(command)

    command = f'java -jar /media/src/Tools/GATK-3.7/GenomeAnalysisTK.jar \
                -T BaseRecalibrator \
                -R /Bioinformatics/01.Reference/b37/b37.fa \
                -L /Bioinformatics/01.Reference/Panel/BCR-ABL1.bed \
                -knownSites /Bioinformatics/01.Reference/b37/a.dbsnp147.b37.vcf \
                -I 03.Output/{name}.indel.bam \
                -o 03.Output/{name}.grp'
    os.system(command)

    command = f'java -jar /media/src/Tools/GATK-3.7/GenomeAnalysisTK.jar \
                -T PrintReads \
                -R /Bioinformatics/01.Reference/b37/b37.fa \
                -L /Bioinformatics/01.Reference/Panel/BCR-ABL1.bed \
                --read_filter MappingQualityZero \
                -BQSR 03.Output/{name}.grp \
                -I 03.Output/{name}.indel.bam \
                -o 03.Output/{name}.bam'
    os.system(command)

    command = f'java -jar /media/src/Tools/GATK-3.7/GenomeAnalysisTK.jar \
                -T HaplotypeCaller \
                -R /Bioinformatics/01.Reference/b37/b37.fa \
                -L /Bioinformatics/01.Reference/Panel/BCR-ABL1.bed \
                --dbsnp /Bioinformatics/01.Reference/b37/a.dbsnp147.b37.vcf \
                -I 03.Output/{name}.bam \
                -o 03.Output/{name}.vcf \
                -stand_call_conf 20'
    os.system(command)

    command = f'samtools mpileup \
                -f /Bioinformatics/01.Reference/b37/b37.fa \
                --max-depth 1000000 \
                03.Output/{name}.bam > 03.Output/{name}.mpileup'
    os.system(command)

    command = f'java -jar /Bioinformatics/00.Tools/varscan-2.4.5/VarScan.v2.4.1.jar \
                mpileup2cns 03.Output/{name}.mpileup \
                --min-avg-qual 20 --min-coverage 10 --min-reads2 5 \
                --min-var-freq 0.01 --variants \ --vcf-sample-list {name} \
                --output-vcf 1 > 03.Output/{name}.varscan.vcf'
    os.system(command)

    # command = f'java -jar /media/src/Tools/GATK-3.7/GenomeAnalysisTK.jar \
    #             -T SelectVariants \
    #             -R /Bioinformatics/01.Reference/b37/b37.fa \
    #             -selectType SNP \
    #             -V 03.Output/{name}.vcf \
    #             -o 03.Output/{name}.SNP.vcf'
    # os.system(command)

    # command = f'java -jar /media/src/Tools/GATK-3.7/GenomeAnalysisTK.jar \
    #             -T SelectVariants \
    #             -R /Bioinformatics/01.Reference/b37/b37.fa \
    #             -selectType INDEL \
    #             -V 03.Output/{name}.vcf \
    #             -o 03.Output/{name}.INDEL.vcf'
    # os.system(command)

    # command = f'java -jar /media/src/Tools/GATK-3.7/GenomeAnalysisTK.jar \
    #             -T VariantFiltration \
    #             -R /Bioinformatics/01.Reference/b37/b37.fa \
    #             -filterName "PASS" \
    #             -filter "QD < 2.0 || FS > 60.0 || MQ < 40.0" \
    #             -V 03.Output/{name}.SNP.vcf \
    #             -o 03.Output/{name}.filtered.SNP.vcf'
    # os.system(command)

    # command = f'java -jar /media/src/Tools/GATK-3.7/GenomeAnalysisTK.jar \
    #             -T VariantFiltration \
    #             -R /Bioinformatics/01.Reference/b37/b37.fa \
    #             -filterName "PASS" \
    #             -filter "QD < 2.0 || FS > 200.0" \
    #             -V 03.Output/{name}.INDEL.vcf \
    #             -o 03.Output/{name}.filtered.INDEL.vcf'
    # os.system(command)

    # command = f'java -jar /Bioinformatics/00.Tools/picard/build/libs/picard.jar \
    #             MergeVcfs \
    #             I=03.Output/{name}.filtered.SNP.vcf \
    #             I=03.Output/{name}.filtered.INDEL.vcf \
    #             O=03.Output/{name}.haplotype.vcf'
    # os.system(command)

    command = f'java -jar /media/src/Tools/GATK-3.7/GenomeAnalysisTK.jar \
               -T CombineVariants \
               -R /Bioinformatics/01.Reference/b37/b37.fa \
               --variant 03.Output/{name}.haplotype.vcf \
               --variant 03.Output/{name}.varscan.vcf \
               -o 03.Output/{name}.final.vcf \
               -genotypeMergeOptions UNIQUIFY'
    # os.system(command)

    command = f'java -jar /media/src/Tools/snpEff/snpEff.jar \
               -v hg{BATCH["Ref.ver"].split("g")[1]} \
               03.Output/{name}.varscan.vcf > 03.Output/{name}.snpeff.vcf'
    os.system(command)
#----------------------------------------------------------------------------------------#
def Annotate(name):
    command = f'convert2annovar.pl -includeinfo -allsample -withfreq -format vcf4 03.Output/{name}.snpeff.vcf > 03.Output/{name}.avinput'
    os.system(command)

    command = f'annotate_variation.pl -geneanno -out 03.Output/{name}.hgvs -build hg{BATCH["Ref.ver"].split("g")[1]} \
                -dbtype refGene \
                -hgvs 03.Output/{name}.avinput /Bioinformatics/00.Tools/annovar/humandb'
    os.system(command)

    command = f'table_annovar.pl 03.Output/{name}.avinput /Bioinformatics/00.Tools/annovar/humandb \
                -buildver hg{BATCH["Ref.ver"].split("g")[1]} -out 03.Output/{name} \
                -remove -protocol \
                refGene,dbnsfp33a,cosmic70,snp138,snp138NonFlagged,popfreq_max_20150413,popfreq_all_20150413,dbscsnv11,exac03nontcga,avsnp147,clinvar_20160302,gnomad_exome,gnomad_genome \
                -operation g,f,f,f,f,f,f,f,f,f,f,f,f \
                -nastring . -otherinfo'
    os.system(command)
# #----------------------------------------------------------------------------------------#
# def Fusion(name, r1, r2):
#     command = f'arriba \
#             -x 03.Output/{name}_Sorted.out.bam \
#             -o 03.Output/{name}_fusions.tsv -O 03.Output/{name}_fusions.discarded.tsv \
#             -a /Bioinformatics/01.Reference/hg{BATCH["Ref.ver"].split("g")[1]}/Homo_sapiens_assembly{BATCH["Ref.ver"].split("g")[1]}.fasta \
#             -g /Bioinformatics/01.Reference/hg{BATCH["Ref.ver"].split("g")[1]}/gencode.v{BATCH["Ref.ver"].split("g")[1]}.annotation.gtf \
#             -b /Bioinformatics/01.Reference/hg{BATCH["Ref.ver"].split("g")[1]}/Arriba/blacklist_hg{BATCH["Ref.ver"].split("g")[1]}_hs37d5_GRCh37_v2.1.0.tsv.gz \
#             -k /Bioinformatics/01.Reference/hg{BATCH["Ref.ver"].split("g")[1]}/Arriba/known_fusions_hg{BATCH["Ref.ver"].split("g")[1]}_hs37d5_GRCh37_v2.1.0.tsv.gz \
#             -p /Bioinformatics/01.Reference/hg{BATCH["Ref.ver"].split("g")[1]}/Arriba/protein_domains_hg{BATCH["Ref.ver"].split("g")[1]}_hs37d5_GRCh37_v2.1.0.gff3'
     #os.system(command)

#     command = f'draw_fusions.R \
#                 --fusions=03.Output/{name}_fusions.tsv \
#                 --alignments=03.Output/{name}_Sorted.out.bam \
#                 --output=03.Output/{name}_fusion.pdf \
#                 --annotation=/Bioinformatics/01.Reference/hg{BATCH["Ref.ver"].split("g")[1]}/gencode.v19.annotation.gtf \
#                 --cytobands=/Bioinformatics/01.Reference/hg{BATCH["Ref.ver"].split("g")[1]}/Arriba/cytobands_hg19_hs37d5_GRCh37_v2.1.0.tsv \
#                 --proteinDomains=/Bioinformatics/01.Reference/hg{BATCH["Ref.ver"].split("g")[1]}/Arriba/protein_domains_hg19_hs37d5_GRCh37_v2.1.0.gff3'
     #os.system(command)
# #----------------------------------------------------------------------------------------#
def MakeSheet(name):
    CS = {}
    RCV = {}
    with open('/Bioinformatics/01.Reference/230211.Clinvar.BCR.ABL1.Variation.txt', 'r') as patho:
        for line in patho:
            line = line.strip()
            splitted = line.split('\t')
            PROTEIN = splitted[2]
            CHARACTER = splitted[5]
            Rcv = splitted[6]
            CS[PROTEIN] = CHARACTER
            RCV[PROTEIN] = Rcv

    DRUG = {}
    with open('/Bioinformatics/01.Reference/230210.ABL.Drug.resistance.txt', 'r') as patho:
        for line in patho:
            line = line.strip()
            splitted = line.split('\t')
            PROTEIN = splitted[0]
            Drug = splitted[1]
            DRUG[PROTEIN] = Drug
    Main_Accession = {'NM_005157': 'O', 'NM_007313' : '', 'NM_004327' : 'O', 'NM_021574' : ''}
    #--------------------------------------------------------------------------------------------------# 
    batch_config = pd.read_csv(f'{Name}.batch.config', sep='\t', header=None, names=['Info'])
    batch_config[['Info', 'value']] = batch_config['Info'].str.split('=', expand=True)
    SAMPLE_COUNT = int(batch_config.iloc[44,1])
    SAMPLE_ID = batch_config.iloc[45,1].split(',')
    SAMPLE_PATH = batch_config.iloc[46,1].split(',')

    COUNT = 0
    BATCH = {}
    for samp, samp_path in zip(SAMPLE_ID, SAMPLE_PATH):
        if os.path.exists(f'{samp_path}03.Output/{samp}.hg19_multianno.txt'):
            COUNT += 1
            with open(f'{samp_path}03.Output/{samp}.hg19_multianno.txt', 'r') as anno:
                for line in anno:
                    if line.startswith('Chr'):
                        continue
                    line = line.strip()
                    splitted = line.split('\t')
                    Batch_info = '_'.join(splitted[0:7])
                    if Batch_info not in BATCH.keys():
                        BATCH[Batch_info] = 1
                    elif Batch_info in BATCH.keys():
                        BATCH[Batch_info] += 1

    if COUNT != SAMPLE_COUNT:
        time.sleep(60)
        MakeSheet(Name)
    elif COUNT == SAMPLE_COUNT:
        pass
    #--------------------------------------------------------------------------------------------------#
    VARIANTS = defaultdict(list)
    Anno = pd.read_csv(f'03.Output/{name}.hg19_multianno.txt', sep='\t')
    Header_Info = list(Anno.columns[10:])
    varscan_format = Anno.iloc[:,143]
    varscan = Anno.iloc[:,144]

    for i in range(Anno.shape[0]):
        Chr = str(Anno.iloc[i,0])
        Start = str(Anno.iloc[i,1])
        End = str(Anno.iloc[i,2])
        Ref = Anno.iloc[i,138]
        Alt = Anno.iloc[i,139]
        Region = Anno.iloc[i,5]
        Gene = Anno.iloc[i,6]
        Other_info = Anno.iloc[i, 10:]
        Gene = Anno.iloc[i,6]
        Flag = Anno.iloc[i,141]

        varscan_per_row = varscan.iloc[i]
        varscan_format_per_row = varscan_format.iloc[i]
        VAF_info = dict(zip(varscan_format_per_row.split(':'), varscan_per_row.split(':')))

        Batch = Anno.iloc[i,0:7].to_list()
        Batch = '_'.join(map(str, Anno.iloc[i,0:7].to_list()))
        #--------------------------------------------------------------------------------------------------#
        if Anno.iloc[i,5] == 'exonic':          
            vtype = Anno.iloc[i,8]
            vinfo = Anno.iloc[i,9]
            for variant in vinfo.split(','):
                variant = variant.split(':')
                NM = variant[1]
                Exon = variant[2]
                HGVSc = variant[3]
                HGVSp = variant[4]
                if HGVSp in CS.keys():
                    Clinical = CS[HGVSp]
                    RCV_clinvar = RCV[HGVSp]
                else:
                    Clinical = 'Uncertain significance'
                    RCV_clinvar = ' '
                if HGVSp in DRUG.keys():
                    Drug = DRUG[HGVSp]
                else:
                    Drug = ' '

                VARIANTS[Chr + Start + Ref + Alt + Gene + HGVSp + NM].extend([
                        ' ', \
                        Main_Accession[NM],\
                        Chr + ':' + Start + '-' + End,\
                        Region,\
                        vtype,\
                        Gene,\
                        NM,\
                        HGVSc,\
                        Ref,\
                        Alt,\
                        HGVSp,\
                        Clinical,\
                        float(VAF_info['FREQ'].replace('%', '')),\
                        int(VAF_info['AD']),\
                        int(VAF_info['DP']),\
                        Flag,\
                        Exon,\
                        RCV_clinvar,\
                        Drug,\
                        str(BATCH[Batch]) + '/' + str(SAMPLE_COUNT)])

                for remain in Other_info:
                    VARIANTS[Chr + Start + Ref + Alt + Gene + HGVSp + NM].append(str(remain))
        #--------------------------------------------------------------------------------------------------#
        elif Region == 'UTR5' or Region == 'UTR3':
            vinfo = Anno.iloc[i,7]
            for variant in vinfo.split(';'):
                variant = variant.split(':')
                NM = variant[0]
                HGVSc = variant[1]
                HGVSp = ' '
                Variant_Type = ' '
                Clinical = ' '
                Exon = ' '
                RCV_clinvar = ' '
                Drug = ' '
                VARIANTS[Chr + Start + Ref + Alt + Gene + HGVSp + NM].extend([
                        ' ', \
                        Main_Accession[NM],\
                        Chr + ':' + Start + '-' + End,\
                        Region,\
                        Variant_Type,\
                        Gene,\
                        NM,\
                        HGVSc,\
                        Ref,\
                        Alt,\
                        HGVSp,\
                        Clinical,\
                        float(VAF_info['FREQ'].replace('%', '')),\
                        int(VAF_info['AD']),\
                        int(VAF_info['DP']),\
                        Flag,\
                        Exon,\
                        RCV_clinvar,\
                        Drug,\
                        str(BATCH[Batch]) + '/' + str(SAMPLE_COUNT)])

                for remain in Other_info:
                    VARIANTS[Chr + Start + Ref + Alt + Gene + HGVSp + NM].append(str(remain))
        #--------------------------------------------------------------------------------------------------#

#Annotation Exel
    Data = pd.DataFrame(VARIANTS)
    Data = Data.transpose()
    COLUMNS = ['Select','Main','Chromosome Position', 'Region','Variant Type', 'Gene', 'NM number', 'HGVSc', 'REF', 'ALT',\
                'HGVSp', 'Clinvar assertion', 'VAF (%)', 'AD', 'DP', \
                'Flag', 'Exon locus', 'RCV.clinvar', 'Drug', 'Same in Batch'] + Header_Info
    Data.columns = COLUMNS

    def text_color(val):
        color = 'red' if not val == 'Uncertain significance' else 'black'
        return 'color: %s' % color

    Data = Data.style.applymap(text_color, subset=pd.IndexSlice[:, ['Clinvar assertion']])

    with pd.ExcelWriter(f'03.Output/{name}.results.xlsx') as excel_note:
        Data.to_excel(excel_note, sheet_name='mutation', index=False)

    # wb = load_workbook(f'03.Output/{name}.results.xlsx')
    # ws = wb['mutation']
    # font = Font(name='Arial')
    # for row in ws.iter_rows():
    #     for cell in row:
    #         cell.font = font
    # wb.save(f'03.Output/{name}.results.xlsx')
#----------------------------------------------------------------------------------------#
def Indexing():
    if os.path.isfile(f'/Bioinformatics/01.Reference/hg{BATCH["Ref.ver"].split("g")[1]}/Homo_sapiens_assembly{BATCH["Ref.ver"].split("g")[1]}.fasta.fai'):
        pass
    else:
        command = f'samtools faidx /Bioinformatics/01.Reference/hg{BATCH["Ref.ver"].split("g")[1]}/Homo_sapiens_assembly{BATCH["Ref.ver"].split("g")[1]}.fasta'
    os.system(command)

    if os.path.isfile(f'/Bioinformatics/01.Reference/hg{BATCH["Ref.ver"].split("g")[1]}/Homo_sapiens_assembly{BATCH["Ref.ver"].split("g")[1]}.dbsnp138.vcf.idx'):
        pass
    else:
        command = f'java -jar /Bioinformatics/00.Tools/gatk-4.1.7.0/gatk-package-4.1.7.0-local.jar \
                    IndexFeatureFile \
                    -I /Bioinformatics/01.Reference/hg{BATCH["Ref.ver"].split("g")[1]}/Homo_sapiens_assembly{BATCH["Ref.ver"].split("g")[1]}.dbsnp138.vcf'
    os.system(command)

    if os.path.isfile(f'/Bioinformatics/01.Reference/hg{BATCH["Ref.ver"].split("g")[1]}/Homo_sapiens_assembly{BATCH["Ref.ver"].split("g")[1]}.dict'):
        pass
    else:
        command = f'java -jar /Bioinformatics/00.Tools/picard/build/libs/picard.jar \
                    CreateSequenceDictionary \
                    R=/Bioinformatics/01.Reference/hg{BATCH["Ref.ver"].split("g")[1]}/Homo_sapiens_assembly{BATCH["Ref.ver"].split("g")[1]}.fasta \
                    O=/Bioinformatics/01.Reference/hg{BATCH["Ref.ver"].split("g")[1]}/Homo_sapiens_assembly{BATCH["Ref.ver"].split("g")[1]}.dict'
    os.system(command)
#----------------------------------------------------------------------------------------#
if BATCH["Step"] == 'All':
    # PreQC(Name, R1, R2)
    # Trimming(Name, R1, R2)
    # PostQC(Name, R1, R2)
    # Refindex()
    # STAR(Name)
    # QC(Name)
    Somatic(Name)
    Annotate(Name)
    MakeSheet(Name)
    # Fusion(Name, R1, R2)
elif BATCH["Step"] == 'FastQC':
    PreQC(Name, R1, R2)
    Trimming(Name, R1, R2)
    PostQC(Name, R1, R2)
elif BATCH["Step"] == 'Trimming':
    Trimming(Name, R1, R2)
elif BATCH["Step"] == 'Align':
    STAR(Name)
elif BATCH["Step"] == 'Mutation':
    Somatic(Name)
# elif BATCH["Step"] == 'Annotation':
#     Annotate(Name)
# elif BATCH["Step"] == 'QC':
#     QC(Name)
# elif BATCH["Step"] == 'Fusion':
#     Fusion(Name, R1, R2)
# elif BATCH["Step"] == 'Results':
#     MakeSheet(Name)
# elif BATCH["Step"] == 'Indexing':
#     Refindex()
